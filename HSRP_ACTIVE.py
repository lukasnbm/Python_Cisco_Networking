# Install "openpyxl" module with bellow command:
# pip install openpyxl

from openpyxl import load_workbook

linelist1  = []
linelist2   = []
linelist3   = []
linelist4   = []


linelist1_1  = []
linelist2_1   = []
linelist3_1   = []
linelist4_1   = []


setlist = []

# Prepare the sheet on excel to work with
wb = load_workbook('C:\Python27\Scripts\HSRP_Active_7600\HSRP_ACTIVE_EXCEL.xlsx')
sheet = wb["HSRP_ACTIVE"]

# Append 1st row of the sheet to linelist1	
for i in range(3, 500):
    linelist1.append(sheet.cell(row=i, column=1).value)

linelist2 = filter (None,linelist1)	 
linelist3 = (' '.join(linelist2)).split() 
#print linelist3

# Add 'P' next to Active|Standby|Init if there are no any of it 
for y in range(len(linelist3)):
    if (linelist3[y] == 'Standby' or linelist3[y] == 'Active' or linelist3[y] == 'Init') and linelist3[y-1] != 'P':
        linelist3.insert(y, 'P')
#print linelist3
		
w=0
e=4
for line in range(len(linelist3)/8):
    linelist4.append(' '.join(linelist3[w:w+1] + linelist3[e:e+1]))
    w = w+8
    e = e+8

#print linelist4	

#############################
#############################

# Append 9th row of the sheet to linelist1_1	
for i in range(3, 500):
    linelist1_1.append(sheet.cell(row=i, column=9).value)

linelist2_1 = filter (None,linelist1_1)	 
linelist3_1 = (' '.join(linelist2_1)).split() 
#print linelist3_1

#Add 'P' next to Active|Standby|Init if there are no any of it 
for y in range(len(linelist3_1)):
    if (linelist3_1[y] == 'Standby' or linelist3_1[y] == 'Active' or linelist3_1[y] == 'Init') and linelist3_1[y-1] != 'P':
        linelist3_1.insert(y, 'P')

#print linelist3_1

w=0
e=4
for line in range(len(linelist3_1)/8):
    linelist4_1.append(' '.join(linelist3_1[w:w+1] + linelist3_1[e:e+1]))
    w = w+8
    e = e+8

#print linelist4_1			 

print ("\n")

#Find and print common value on linelist4 and linelist4_1 	

setlist = list(set(linelist4).intersection(linelist4_1))
print "VLAN(S) on ACTIVE-ACTIVE Condition"
print "or on INIT-INIT Condition"
print "==================================\n"
for c in range (len(setlist)):
    print setlist[c]	

	
raw_input()	