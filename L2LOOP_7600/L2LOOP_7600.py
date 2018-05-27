# Install "openpyxl" module with bellow command:
# pip install openpyxl

from openpyxl import load_workbook

linelist1  = []
linelistfinal1  = []
linelist2   = []
linelistfinal2  = []
setlist = []

#Prepare the sheet on excel to work with
wb = load_workbook('C:\Python27\Scripts\L2LOOP 7600\L2EXCEL.xlsx')
sheet = wb["L2_Sheet1"]

# Append 1st row of the sheet to linelistfinal1	
for i in range(2, 500):
    linelist1.append(sheet.cell(row=i, column=1).value)

linelistfinal1 = filter (None,linelist1)	 
#print linelistfinal1		 

print ("\n")

# Append 2nd row of the sheet to linelist2
for i in range(2, 500):
    linelist2.append(sheet.cell(row=i, column=2).value)

linelistfinal2 = filter (None,linelist2)	 
#print linelistfinal2

#Find and print common value on linelist1 and linelist2 	
setlist = list(set(linelistfinal1).intersection(linelistfinal2))
print "COMMON VLAN(S)"
print "=============\n"
for c in range (len(setlist)):
    print setlist[c]	
	
raw_input()	
