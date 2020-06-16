# Install "openpyxl" module with bellow command:
# pip install openpyxl
# Author: Lukas Marbun
# Dimension Data Indonesia 2019

from openpyxl import load_workbook

#Prepare the sheet on excel to work with
wb = load_workbook('BSC_Integration.xlsx')
sheet_aoip = wb["AOIP"]
sheet_gboip = wb["GBOIP"]
sheet_abis = wb["ABIS"]
sheet_oam = wb["OAM"]

#Create the new file
f= open("BSC_INTEGRATION_SCRIPT_XR.txt","w+")

row_number_aoip = sheet_aoip.max_row
row_number_gboip = sheet_gboip.max_row
row_number_abis  = sheet_abis.max_row
row_number_oam  = sheet_oam.max_row

print ("## Enter router name  ##")
name1 = raw_input('Router 1 Name : ')
name2 = raw_input('Router 2 Name : ')
print ("\n")

phy_int_1 = []
bdi_1 = []
phy_int_2 = []
bdi_2 = []


##################
#  AOIP LOOPING  #
##################

cells_aoip1 = sheet_aoip['A3':'J'+ str(row_number_aoip)]

print ("ROUTER 1 AOIP")
print ("====DONE=====\n")

f.write ("==============\n")
f.write ("|    AOIP    |\n")
f.write ("==============\n\n")
f.write ("%s \n" %name1)
f.write ("================\n")


for c1,c2,c3,c4,c5,c6,c7,c8,c9,c10 in cells_aoip1:
    f.write ("interface {0}.{1}\n".format(c2.value,c1.value))
    f.write ("interface {0}.{1} description {2}\n".format(c2.value,c1.value,c3.value))
    f.write ("interface {0}.{1} vrf {2}\n".format(c2.value,c1.value,c4.value))
    f.write ("interface {0}.{1} ipv4 address {2} {3}\n".format(c2.value,c1.value,c5.value,c6.value))
    f.write ("interface {0}.{1} encapsulation dot1q {2}\n\n".format(c2.value,c1.value,c1.value))
    f.write ("router static vrf {0} address-family ipv4 unicast {1} {2} {3}.{4} {5} bfd fast-detect minimum-interval 200 multiplier 3 tag {6} \n".format(c4.value,c8.value,c9.value,c2.value,c1.value,c7.value,c10.value))
    f.write ("\n\n")

cells_aoip2 = sheet_aoip['K3':'T'+ str(row_number_aoip)]

print ("ROUTER 2 AOIP")
print ("====DONE=====\n")

f.write ("%s \n" %name2)
f.write ("================\n")

for c1,c2,c3,c4,c5,c6,c7,c8,c9,c10 in cells_aoip2:
    f.write ("interface {0}.{1}\n".format(c2.value,c1.value))
    f.write ("interface {0}.{1} description {2}\n".format(c2.value,c1.value,c3.value))
    f.write ("interface {0}.{1} vrf {2}\n".format(c2.value,c1.value,c4.value))
    f.write ("interface {0}.{1} ipv4 address {2} {3}\n".format(c2.value, c1.value, c5.value, c6.value))
    f.write ("interface {0}.{1} encapsulation dot1q {2}\n\n".format(c2.value,c1.value,c1.value))
    f.write ("router static vrf {0} address-family ipv4 unicast {1} {2} {3}.{4} {5} bfd fast-detect minimum-interval 200 multiplier 3 tag {6} \n".format(c4.value,c8.value,c9.value,c2.value,c1.value,c7.value,c10.value))
    f.write ("\n\n")


###################
#  GBOIP LOOPING  #
###################

cells_gboip1 = sheet_gboip['A3':'J'+ str(row_number_gboip)]

print ("ROUTER 1 GBOIP")
print ("=====DONE=====\n")

f.write ("==============\n")
f.write ("|    GBOIP   |\n")
f.write ("==============\n\n")
f.write ("%s \n" %name1)
f.write ("================\n")


for c1,c2,c3,c4,c5,c6,c7,c8,c9,c10 in cells_gboip1:
    f.write ("interface BVI{0}\n".format(c1.value))
    f.write ("interface BVI{0} description {1}\n".format(c1.value,c3.value))
    f.write ("interface BVI{0} vrf {1}\n".format(c1.value,c4.value))
    f.write ("interface BVI{0} ipv4 address {1} {2} \n\n".format(c1.value,c5.value,c6.value))
    f.write ("interface {0} description {1}\n".format(c2.value,c3.value))
    f.write ("interface {0} \n".format(c2.value))
    f.write ("interface {0} negotiation auto \n".format(c2.value))
    f.write ("interface {0} l2transport \n\n".format(c2.value))
    f.write ("l2vpn bridge group BVI bridge-domain {0} \n".format(c1.value))
    f.write ("l2vpn bridge group BVI bridge-domain {0} interface {1} \n".format(c1.value,c2.value))
    f.write ("l2vpn bridge group BVI bridge-domain {0} routed interface BVI{1} \n\n".format(c1.value,c1.value))
    f.write ("router static vrf {0} address-family ipv4 unicast {1} {2} BVI{3} {4} bfd fast-detect minimum-interval 200 multiplier 3 tag {5} \n".format(c4.value,c8.value,c9.value,c1.value,c7.value,c10.value))
    f.write ("\n")


cells_gboip2 = sheet_gboip['K3':'T'+ str(row_number_gboip)]

print ("ROUTER 2 GBOIP")
print ("=====DONE=====\n")

f.write ("%s \n" %name2)
f.write ("================\n")

for c1,c2,c3,c4,c5,c6,c7,c8,c9,c10 in cells_gboip2:
    f.write ("interface BVI{0}\n".format(c1.value))
    f.write ("interface BVI{0} description {1}\n".format(c1.value,c3.value))
    f.write ("interface BVI{0} vrf {1} \n".format(c1.value,c4.value))
    f.write ("interface BVI{0} ipv4 address {1} {2} \n\n".format(c1.value,c5.value,c6.value))
    f.write ("interface {0} description {1}\n".format(c2.value,c3.value))
    f.write ("interface {0} \n".format(c2.value))
    f.write ("interface {0} negotiation auto \n".format(c2.value))
    f.write ("interface {0} l2transport \n\n".format(c2.value))
    f.write ("l2vpn bridge group BVI bridge-domain {0} \n".format(c1.value))
    f.write ("l2vpn bridge group BVI bridge-domain {0} interface {1} \n".format(c1.value,c2.value))
    f.write ("l2vpn bridge group BVI bridge-domain {0} routed interface BVI{1} \n\n".format(c1.value,c1.value))
    f.write ("router static vrf {0} address-family ipv4 unicast {1} {2} BVI{3} {4} bfd fast-detect minimum-interval 200 multiplier 3 tag {5} \n".format(c4.value,c8.value,c9.value,c1.value,c7.value,c10.value))
    f.write ("\n")


##################
#  ABIS LOOPING  #
##################

cells_abis1 = sheet_abis['A3':'J'+ str(row_number_abis)]

print ("ROUTER 1 ABIS")
print ("=====DONE=====\n")

f.write ("==============\n")
f.write ("|    ABIS    |\n")
f.write ("==============\n\n")
f.write ("%s \n" %name1)
f.write ("================\n")


for c1,c2,c3,c4,c5,c6,c7,c8,c9,c10 in cells_abis1:
    f.write ("interface BVI{0}\n".format(c1.value))
    f.write ("interface BVI{0} description {1}\n".format(c1.value,c3.value))
    f.write ("interface BVI{0} vrf {1}\n".format(c1.value,c4.value))
    f.write ("interface BVI{0} ipv4 address {1} {2} \n\n".format(c1.value,c5.value,c6.value))
    f.write ("interface {0} description {1}\n".format(c2.value,c3.value))
    f.write ("interface {0} \n".format(c2.value))
    f.write ("interface {0} negotiation auto \n".format(c2.value))
    f.write ("interface {0} l2transport \n\n".format(c2.value))
    f.write ("l2vpn bridge group BVI bridge-domain {0} \n".format(c1.value))
    f.write ("l2vpn bridge group BVI bridge-domain {0} interface {1} \n".format(c1.value,c2.value))
    f.write ("l2vpn bridge group BVI bridge-domain {0} routed interface BVI{1} \n\n".format(c1.value,c1.value))
    f.write ("router static vrf {0} address-family ipv4 unicast {1} {2} BVI{3} {4} bfd fast-detect minimum-interval 200 multiplier 3 tag {5} \n".format(c4.value,c8.value,c9.value,c1.value,c7.value,c10.value))
    f.write ("\n")

cells_abis2 = sheet_abis['K3':'T'+ str(row_number_abis)]

print ("ROUTER 2 ABIS")
print ("=====DONE=====\n")

f.write ("%s \n" %name2)
f.write ("================\n")

for c1,c2,c3,c4,c5,c6,c7,c8,c9,c10 in cells_abis2:
    f.write ("interface BVI{0}\n".format(c1.value))
    f.write ("interface BVI{0} description {1}\n".format(c1.value,c3.value))
    f.write ("interface BVI{0} vrf {1}\n".format(c1.value,c4.value))
    f.write ("interface BVI{0} ipv4 address {1} {2} \n\n".format(c1.value,c5.value,c6.value))
    f.write ("interface {0} description {1}\n".format(c2.value,c3.value))
    f.write ("interface {0} \n".format(c2.value))
    f.write ("interface {0} negotiation auto \n".format(c2.value))
    f.write ("interface {0} l2transport \n\n".format(c2.value))
    f.write ("l2vpn bridge group BVI bridge-domain {0} \n".format(c1.value))
    f.write ("l2vpn bridge group BVI bridge-domain {0} interface {1} \n".format(c1.value,c2.value))
    f.write ("l2vpn bridge group BVI bridge-domain {0} routed interface BVI{1} \n\n".format(c1.value,c1.value))
    f.write ("router static vrf {0} address-family ipv4 unicast {1} {2} BVI{3} {4} bfd fast-detect minimum-interval 200 multiplier 3 tag {5} \n".format(c4.value,c8.value,c9.value,c1.value,c7.value,c10.value))
    f.write ("\n")



##################
#   OAM LOOPING  #
##################

cells_oam1 = sheet_oam['A3':'I'+ str(row_number_oam)]

print ("ROUTER 1 OAM")
print ("====DONE====\n")

f.write ("==============\n")
f.write ("|    OAM     |\n")
f.write ("==============\n\n")
f.write ("%s \n" %name1)
f.write ("================\n\n")
f.write ("!!NOTES : Change interface XYZ to L2 Interlink Interface. Ex: BE10, Gi0/4/0/4, etc  \n\n")

for c1,c2,c3,c4,c5,c6,c7,c8,c9 in cells_oam1:
    f.write ("interface XYZ.{0} l2transport\n".format(c1.value))
    f.write ("interface XYZ.{0} l2transport encapsulation dot1q {1} exact \n".format(c1.value,c1.value))
    f.write ("interface XYZ.{0} l2transport rewrite ingress tag pop 1 symmetric\n\n".format(c1.value))
    f.write ("interface BVI{0}\n".format(c1.value))
    f.write ("interface BVI{0} description {1}\n".format(c1.value,c3.value))
    f.write ("interface BVI{0} vrf {1}\n".format(c1.value,c4.value))
    f.write ("interface BVI{0} ipv4 address {1} {2} \n\n".format(c1.value,c5.value,c6.value))
    f.write ("interface {0} description {1}\n".format(c2.value,c3.value))
    f.write ("interface {0} \n".format(c2.value))
    f.write ("interface {0} negotiation auto \n".format(c2.value))
    f.write ("interface {0} load-interval 30 \n".format(c2.value))
    f.write ("interface {0} l2transport \n\n".format(c2.value))
    f.write ("l2vpn bridge group BVI bridge-domain {0} \n".format(c1.value))
    f.write ("l2vpn bridge group BVI bridge-domain {0} interface {1} \n".format(c1.value,c2.value))
    f.write ("l2vpn bridge group BVI bridge-domain {0} routed interface BVI{1} \n".format(c1.value,c1.value))
    f.write ("l2vpn bridge group BVI bridge-domain {0} interface XYZ.{0}\n\n".format(c1.value))
    f.write ("router hsrp interface BVI.{0} \n".format(c1.value))
    f.write ("router hsrp interface BVI.{0} address-family ipv4 hsrp {1}\n".format(c1.value,c8.value))
    f.write ("router hsrp interface BVI.{0} address-family ipv4 hsrp {1} preempt \n".format(c1.value,c8.value))
    f.write ("router hsrp interface BVI.{0} address-family ipv4 hsrp {1} priority 110 \n".format(c1.value,c8.value))
    f.write ("router hsrp interface BVI.{0} address-family ipv4 hsrp {1} address {2} \n".format(c1.value,c8.value,c7.value))
    f.write ("router hsrp interface BVI.{0} address-family ipv4 hsrp version 2 \n\n".format(c1.value,c8.value))
    f.write ("\n")


cells_oam2 = sheet_oam['J3':'R'+ str(row_number_oam)]

print ("ROUTER 2 OAM")
print ("====DONE====\n")

f.write ("%s \n" %name2)
f.write ("================\n\n")
f.write ("!!NOTES : Change interface XYZ to L2 Interlink Interface. Ex: BE10, Gi0/4/0/4, etc  \n\n")

for c1,c2,c3,c4,c5,c6,c7,c8,c9 in cells_oam2:
    f.write ("interface XYZ.{0} l2transport\n".format(c1.value))
    f.write ("interface XYZ.{0} l2transport encapsulation dot1q {1} exact \n".format(c1.value,c1.value))
    f.write ("interface XYZ.{0} l2transport rewrite ingress tag pop 1 symmetric\n\n".format(c1.value))
    f.write ("interface BVI{0}\n".format(c1.value))
    f.write ("interface BVI{0} description {1}\n".format(c1.value,c3.value))
    f.write ("interface BVI{0} vrf {1}\n".format(c1.value,c4.value))
    f.write ("interface BVI{0} ipv4 address {1} {2} \n\n".format(c1.value,c5.value,c6.value))
    f.write ("interface {0} description {1}\n".format(c2.value,c3.value))
    f.write ("interface {0} \n".format(c2.value))
    f.write ("interface {0} negotiation auto \n".format(c2.value))
    f.write ("interface {0} load-interval 30 \n".format(c2.value))
    f.write ("interface {0} l2transport \n\n".format(c2.value))
    f.write ("l2vpn bridge group BVI bridge-domain {0} \n".format(c1.value))
    f.write ("l2vpn bridge group BVI bridge-domain {0} interface {1} \n".format(c1.value,c2.value))
    f.write ("l2vpn bridge group BVI bridge-domain {0} routed interface BVI{1} \n".format(c1.value,c1.value))
    f.write ("l2vpn bridge group BVI bridge-domain {0} interface XYZ.{0}\n\n".format(c1.value))
    f.write ("router hsrp interface BVI.{0} \n".format(c1.value))
    f.write ("router hsrp interface BVI.{0} address-family ipv4 hsrp {1}\n".format(c1.value,c8.value))
    f.write ("router hsrp interface BVI.{0} address-family ipv4 hsrp {1} preempt \n".format(c1.value,c8.value))
    f.write ("router hsrp interface BVI.{0} address-family ipv4 hsrp {1} priority 90 \n".format(c1.value,c8.value))
    f.write ("router hsrp interface BVI.{0} address-family ipv4 hsrp {1} address {2} \n".format(c1.value,c8.value,c7.value))
    f.write ("router hsrp interface BVI.{0} address-family ipv4 hsrp version 2 \n\n".format(c1.value,c8.value))
    f.write ("\n")

f.close()

raw_input()