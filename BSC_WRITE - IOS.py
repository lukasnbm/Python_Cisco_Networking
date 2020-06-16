# Install "openpyxl" module with bellow command:
# pip install openpyxl
# Author: Lukas Marbun
# Dimension Data Indonesia 2019

from openpyxl import load_workbook

#Prepare the sheet on excel to work with
wb = load_workbook('BSC_Integration.xlsx')
sheet_aoip = wb["AOIP"]
sheet_gboip = wb["GBOIP"]
sheet_abis = wb["ABIS"]
sheet_oam = wb["OAM"]

#Create the new file
f= open("BSC_INTEGRATION_SCRIPT.txt","w+")

row_number_aoip = sheet_aoip.max_row
row_number_gboip = sheet_gboip.max_row
row_number_abis  = sheet_abis.max_row
row_number_oam  = sheet_oam.max_row

print ("## Enter router name  ##")
name1 = raw_input('Router 1 Name : ')
name2 = raw_input('Router 2 Name : ')
print ("\n")

phy_int_1 = []
bdi_1 = []
phy_int_2 = []
bdi_2 = []


##################
#  AOIP LOOPING  #
##################

cells_aoip1 = sheet_aoip['A3':'J'+ str(row_number_aoip)]

print ("ROUTER 1 AOIP")
print ("====DONE=====\n")

f.write ("==============\n")
f.write ("|    AOIP    |\n")
f.write ("==============\n\n")
f.write ("%s \n" %name1)
f.write ("================\n")


for c1,c2,c3,c4,c5,c6,c7,c8,c9,c10 in cells_aoip1:
    phy_int_1.append(c2.value)
    bdi_1.append(c1.value)
    f.write ("interface BDI{0}\n".format(c1.value))
    f.write (" description {0}\n".format(c3.value))
    f.write (" ip vrf forwarding {0}\n".format(c4.value))
    f.write (" ip address {0} {1}\n".format(c5.value,c6.value))
    f.write (" no ip redirects\n")
    f.write (" no ip proxy-arp\n")
    f.write (" bfd interval 200 min_rx 200 multiplier 3\n\n")
    f.write ("ip route static bfd BDI{0} {1}\n".format(c1.value,c7.value))
    f.write ("ip route vrf {0} {1} {2} BDI{3} {4} tag {5} name {6}\n".format(c4.value,c8.value,c9.value,c1.value,c7.value,c10.value,c3.value))
    f.write ("\n")

List_length = (len(phy_int_1))
x=0
y=0

for i in range(0,List_length,2):
    f.write ("interface " + phy_int_1[i] + "\n")
    f.write (" description AoIP_1_BSC_\n")
    f.write (" no ip address\n")
    f.write (" load-interval 30\n")
    f.write (" negotiation auto\n")
    f.write (" service instance trunk " + str(bdi_1[i]) + " ethernet\n")
    f.write ("  encapsulation dot1q ")
    for x in range(y,y+2):
        f.write (str(bdi_1[x]) + ",")
    y+=2
    f.write ("\b\n")
    f.write ("  rewrite ingress tag pop 1 symmetric\n")
    f.write ("  bridge-domain from-encapsulation\n\n\n")


cells_aoip2 = sheet_aoip['K3':'T'+ str(row_number_aoip)]

print ("ROUTER 2 AOIP")
print ("====DONE=====\n")

f.write ("%s \n" %name2)
f.write ("================\n")

for c1,c2,c3,c4,c5,c6,c7,c8,c9,c10 in cells_aoip2:
    phy_int_2.append(c2.value)
    bdi_2.append(c1.value)
    f.write ("interface BDI{0}\n".format(c1.value))
    f.write (" description {0}\n".format(c3.value))
    f.write (" ip vrf forwarding {0}\n".format(c4.value))
    f.write (" ip address {0} {1}\n".format(c5.value,c6.value))
    f.write (" no ip redirects\n")
    f.write (" no ip proxy-arp\n")
    f.write (" bfd interval 200 min_rx 200 multiplier 3\n\n")
    f.write ("ip route static bfd BDI{0} {1}\n".format(c1.value,c7.value))
    f.write ("ip route vrf {0} {1} {2} BDI{3} {4} tag {5} name {6}\n".format(c4.value,c8.value,c9.value,c1.value,c7.value,c10.value,c3.value))
    f.write ("\n")

List_length = (len(phy_int_2))
x=0
y=0

for i in range(0,List_length,2):
    f.write ("interface " + phy_int_2[i] + "\n")
    f.write (" description AoIP_2_BSC_\n")
    f.write (" no ip address\n")
    f.write (" load-interval 30\n")
    f.write (" negotiation auto\n")
    f.write (" service instance trunk " + str(bdi_2[i]) + " ethernet\n")
    f.write ("  encapsulation dot1q ")
    for x in range(y,y+2):
        f.write (str(bdi_2[x]) + ",")
    y+=2
    f.write ("\b\n")
    f.write ("  rewrite ingress tag pop 1 symmetric\n")
    f.write ("  bridge-domain from-encapsulation\n\n\n")

del phy_int_1[:]
del bdi_1[:]
del phy_int_2[:]
del bdi_2[:]


###################
#  GBOIP LOOPING  #
###################

cells_gboip1 = sheet_gboip['A3':'J'+ str(row_number_gboip)]

print ("ROUTER 1 GBOIP")
print ("=====DONE=====\n")

f.write ("==============\n")
f.write ("|    GBOIP   |\n")
f.write ("==============\n\n")
f.write ("%s \n" %name1)
f.write ("================\n")


for c1,c2,c3,c4,c5,c6,c7,c8,c9,c10 in cells_gboip1:
    phy_int_1.append(c2.value)
    bdi_1.append(c1.value)
    f.write ("interface BDI{0}\n".format(c1.value))
    f.write (" description {0}\n".format(c3.value))
    f.write (" ip vrf forwarding {0}\n".format(c4.value))
    f.write (" ip address {0} {1}\n".format(c5.value,c6.value))
    f.write (" no ip redirects\n")
    f.write (" no ip proxy-arp\n")
    f.write (" bfd interval 200 min_rx 200 multiplier 3\n\n")
    f.write ("ip route static bfd BDI{0} {1}\n".format(c1.value,c7.value))
    f.write ("ip route vrf {0} {1} {2} BDI{3} {4} tag {5} name {6}\n".format(c4.value,c8.value,c9.value,c1.value,c7.value,c10.value,c3.value))
    f.write ("\n")

List_length = (len(phy_int_1))
x=0
y=0

for i in range(0,List_length):
    f.write ("interface " + phy_int_1[i] + "\n")
    f.write (" description GboIP_1_BSC_\n")
    f.write (" no ip address\n")
    f.write (" load-interval 30\n")
    f.write (" negotiation auto\n")
    f.write (" service instance " + str(bdi_1[i]) + " ethernet\n")
    f.write ("  encapsulation untagged\n")
    for x in range(y,y+1):
        f.write ("  bridge-domain " + str(bdi_1[x]))
    y+=1
    f.write ("\n\n")


cells_gboip2 = sheet_gboip['K3':'T'+ str(row_number_gboip)]

print ("ROUTER 2 GBOIP")
print ("=====DONE=====\n")

f.write ("%s \n" %name2)
f.write ("================\n")

for c1,c2,c3,c4,c5,c6,c7,c8,c9,c10 in cells_gboip2:
    phy_int_2.append(c2.value)
    bdi_2.append(c1.value)
    f.write ("interface BDI{0}\n".format(c1.value))
    f.write (" description {0}\n".format(c3.value))
    f.write (" ip vrf forwarding {0}\n".format(c4.value))
    f.write (" ip address {0} {1}\n".format(c5.value,c6.value))
    f.write (" no ip redirects\n")
    f.write (" no ip proxy-arp\n")
    f.write (" bfd interval 200 min_rx 200 multiplier 3\n\n")
    f.write ("ip route static bfd BDI{0} {1}\n".format(c1.value,c7.value))
    f.write ("ip route vrf {0} {1} {2} BDI{3} {4} tag {5} name {6}\n".format(c4.value,c8.value,c9.value,c1.value,c7.value,c10.value,c3.value))
    f.write ("\n")

List_length = (len(phy_int_2))
x=0
y=0

for i in range(0,List_length):
    f.write ("interface " + phy_int_2[i] + "\n")
    f.write (" description GboIP_2_BSC_\n")
    f.write (" no ip address\n")
    f.write (" load-interval 30\n")
    f.write (" negotiation auto\n")
    f.write (" service instance " + str(bdi_2[i]) + " ethernet\n")
    f.write ("  encapsulation untagged\n")
    for x in range(y,y+1):
        f.write ("  bridge-domain " + str(bdi_2[x]))
    y+=1
    f.write ("\n\n")

del phy_int_1[:]
del bdi_1[:]
del phy_int_2[:]
del bdi_2[:]


##################
#  ABIS LOOPING  #
##################

cells_abis1 = sheet_abis['A3':'J'+ str(row_number_abis)]

print ("ROUTER 1 ABIS")
print ("=====DONE=====\n")

f.write ("==============\n")
f.write ("|    ABIS    |\n")
f.write ("==============\n\n")
f.write ("%s \n" %name1)
f.write ("================\n")


for c1,c2,c3,c4,c5,c6,c7,c8,c9,c10 in cells_abis1:
    phy_int_1.append(c2.value)
    bdi_1.append(c1.value)
    f.write ("interface BDI{0}\n".format(c1.value))
    f.write (" description {0}\n".format(c3.value))
    f.write (" ip vrf forwarding {0}\n".format(c4.value))
    f.write (" ip address {0} {1}\n".format(c5.value,c6.value))
    f.write (" no ip redirects\n")
    f.write (" no ip proxy-arp\n")
    f.write (" bfd interval 200 min_rx 200 multiplier 3\n\n")
    f.write ("ip route static bfd BDI{0} {1}\n".format(c1.value,c7.value))
    f.write ("ip route vrf {0} {1} {2} BDI{3} {4} tag {5} name {6}\n".format(c4.value,c8.value,c9.value,c1.value,c7.value,c10.value,c3.value))
    f.write ("\n")

List_length = (len(phy_int_1))
x=0
y=0

for i in range(0,List_length):
    f.write ("interface " + phy_int_1[i] + "\n")
    f.write (" description Abis_1_BSC_\n")
    f.write (" no ip address\n")
    f.write (" load-interval 30\n")
    f.write (" negotiation auto\n")
    f.write (" service instance " + str(bdi_1[i]) + " ethernet\n")
    f.write ("  encapsulation untagged\n")
    for x in range(y,y+1):
        f.write ("  bridge-domain " + str(bdi_1[x]))
    y+=1
    f.write ("\n\n")


cells_abis2 = sheet_abis['K3':'T'+ str(row_number_abis)]

print ("ROUTER 2 ABIS")
print ("=====DONE=====\n")

f.write ("%s \n" %name2)
f.write ("================\n")

for c1,c2,c3,c4,c5,c6,c7,c8,c9,c10 in cells_abis2:
    phy_int_2.append(c2.value)
    bdi_2.append(c1.value)
    f.write ("interface BDI{0}\n".format(c1.value))
    f.write (" description {0}\n".format(c3.value))
    f.write (" ip vrf forwarding {0}\n".format(c4.value))
    f.write (" ip address {0} {1}\n".format(c5.value,c6.value))
    f.write (" no ip redirects\n")
    f.write (" no ip proxy-arp\n")
    f.write (" bfd interval 200 min_rx 200 multiplier 3\n\n")
    f.write ("ip route static bfd BDI{0} {1}\n".format(c1.value,c7.value))
    f.write ("ip route vrf {0} {1} {2} BDI{3} {4} tag {5} name {6}\n".format(c4.value,c8.value,c9.value,c1.value,c7.value,c10.value,c3.value))
    f.write ("\n")

List_length = (len(phy_int_2))
x=0
y=0

for i in range(0,List_length):
    f.write ("interface " + phy_int_2[i] + "\n")
    f.write (" description Abis_2_BSC_\n")
    f.write (" no ip address\n")
    f.write (" load-interval 30\n")
    f.write (" negotiation auto\n")
    f.write (" service instance " + str(bdi_2[i]) + " ethernet\n")
    f.write ("  encapsulation untagged\n")
    for x in range(y,y+1):
        f.write ("  bridge-domain " + str(bdi_2[x]))
    y+=1
    f.write ("\n\n")

del phy_int_1[:]
del bdi_1[:]
del phy_int_2[:]
del bdi_2[:]


##################
#   OAM LOOPING  #
##################

cells_oam1 = sheet_oam['A3':'I'+ str(row_number_oam)]

print ("ROUTER 1 OAM")
print ("====DONE====\n")

f.write ("==============\n")
f.write ("|    OAM     |\n")
f.write ("==============\n\n")
f.write ("%s \n" %name1)
f.write ("================\n\n")
f.write ("!!NOTES : Change interface XYZ to L2 Interlink Interface. Ex: BE10, Gi0/4/0/4, etc  \n\n")


for c1,c2,c3,c4,c5,c6,c7,c8,c9 in cells_oam1:
    phy_int_1.append(c2.value)
    bdi_1.append(c1.value)
    f.write ("interface BDI{0}\n".format(c1.value))
    f.write (" description {0}\n".format(c3.value))
    f.write (" ip vrf forwarding {0}\n".format(c4.value))
    f.write (" ip address {0} {1}\n".format(c5.value,c6.value))
    f.write (" no ip redirects\n")
    f.write (" no ip proxy-arp\n")
    f.write (" standby {0} ip {1}\n".format(c8.value,c7.value))
    f.write (" standby {0} priority 110\n".format(c8.value))
    f.write (" standby {0} preempt\n".format(c8.value))
    f.write ("\n")

List_length = (len(phy_int_1))
x=0
y=0

for i in range(0,List_length):
    f.write ("interface " + phy_int_1[i] + "\n")
    f.write (" description OAM_1_BSC_\n")
    f.write (" no ip address\n")
    f.write (" load-interval 30\n")
    f.write (" negotiation auto\n")
    f.write (" service instance " + str(bdi_1[i]) + " ethernet\n")
    f.write ("  encapsulation untagged\n")
    for x in range(y,y+1):
        f.write ("  bridge-domain " + str(bdi_1[x]))
    f.write ("\n\n")
    f.write("interface XYZ  <-- L2 Interlink Interface" + "\n")
    f.write(" service instance " + str(bdi_1[i]) + " ethernet\n")
    f.write("  encapsulation dot1q ")
    for x in range(y,y+1):
        f.write (str(bdi_1[x])+ "\n")
    f.write("  rewrite ingress tag pop 1 symmetric\n")
    for x in range(y,y+1):
        f.write ("  bridge-domain " + str(bdi_1[x]))
    y += 1
    f.write("\n\n")


cells_oam2 = sheet_oam['J3':'R'+ str(row_number_oam)]

print ("ROUTER 2 OAM")
print ("====DONE====\n")

f.write ("%s \n" %name2)
f.write ("================\n\n")
f.write ("!!NOTES : Change interface XYZ to L2 Interlink Interface. Ex: BE10, Gi0/4/0/4, etc  \n\n")

for c1,c2,c3,c4,c5,c6,c7,c8,c9 in cells_oam2:
    phy_int_2.append(c2.value)
    bdi_2.append(c1.value)
    f.write ("interface BDI{0}\n".format(c1.value))
    f.write (" description {0}\n".format(c3.value))
    f.write (" ip vrf forwarding {0}\n".format(c4.value))
    f.write (" ip address {0} {1}\n".format(c5.value,c6.value))
    f.write (" no ip redirects\n")
    f.write (" no ip proxy-arp\n")
    f.write (" standby {0} ip {1}\n".format(c8.value,c7.value))
    f.write (" standby {0} priority 90\n".format(c8.value))
    f.write (" standby {0} preempt\n".format(c8.value))
    f.write ("\n")

List_length = (len(phy_int_2))
x=0
y=0

for i in range(0,List_length):
    f.write ("interface " + phy_int_2[i] + "\n")
    f.write (" description OAM_2_BSC_\n")
    f.write (" no ip address\n")
    f.write (" load-interval 30\n")
    f.write (" negotiation auto\n")
    f.write (" service instance " + str(bdi_2[i]) + " ethernet\n")
    f.write ("  encapsulation untagged\n")
    for x in range(y,y+1):
        f.write ("  bridge-domain " + str(bdi_2[x]))
    f.write ("\n\n")
    f.write("interface XYZ  <-- L2 Interlink Interface" + "\n")
    f.write(" service instance " + str(bdi_2[i]) + " ethernet\n")
    f.write("  encapsulation dot1q ")
    for x in range(y,y+1):
        f.write (str(bdi_2[x])+ "\n")
    f.write("  rewrite ingress tag pop 1 symmetric\n")
    for x in range(y,y+1):
        f.write ("  bridge-domain " + str(bdi_1[x]))
    y += 1
    f.write("\n\n")

del phy_int_1[:]
del bdi_1[:]
del phy_int_2[:]
del bdi_2[:]


f.close()

raw_input()